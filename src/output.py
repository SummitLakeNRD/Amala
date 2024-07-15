import json
import cv2
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from math import ceil
import numpy as np

class textOut:
    def __init__(self):
        self.json = {}
        self.dateFormat = "%Y:%m:%d %H:%M:%S"
        self.outputDir = os.path.join(os.getcwd(), "output", "jsons")
        os.makedirs(self.outputDir, exist_ok=True)
        self.counter = 0

    def output(self, exifData, classes, confidence,
               birdCoordinates, yoloCoordinates):
        # Takes in output from neural network (AI) inference and 
        # image EXIF data and returns a json file
        image_ids = []
        for w, x, y, z in zip(birdCoordinates, yoloCoordinates, 
                              classes, confidence):
            dateString = datetime.strptime(exifData['dateTime'], self.dateFormat)
            fileName = "duckData_" + str(self.counter) + ".json"
            self.json = {'dateTime': dateString.strftime('%Y-%m-%d %H:%M:%S'),
                         'birdID': self.counter,
                         'class': y, 
                         'confidence': round(z, 3),
                         'species': "",
                         'birdNorthing': round(w[1], 2),
                         'birdEasting': round(w[0], 2),
                         'utmZone': exifData['utmZone'],
                         'imageBearing': exifData['imageBearing'],
                         'flightHeight_m': exifData['flightHeight_m'],
                         'pixelWidth': exifData['imageWidth'],
                         'pixelHeight': exifData['imageHeight'],
                         'center_x': x[0],
                         'center_y': x[1],
                         'box_width': x[2],
                         'box_height': x[3],
                         'filename': exifData['imageFile']}
            image_ids.append(str(self.counter))
            self.counter += 1
            with open(os.path.join(self.outputDir, fileName), "w") as f:
                json.dump(self.json, f)
        return image_ids

class imageOut:
    def __init__(self):
        self.font = cv2.FONT_HERSHEY_SIMPLEX
        self.thickness = 3
        self.fontColor = (43, 75, 238)
        self.fontSize = 1
        self.outputDir = os.path.join(os.getcwd(), "output")
        self.image_counter = 0

    def output(self, frame, bbox, image_ids):
        # Generate output image that will be used for supervised
        # species identification of positively identified waterfowl
        if len(bbox) == 0:
            return 0 
        else:
            imageName = "duckImage_" + str(self.image_counter) + ".png"
            centers = [[ceil((i[2] + i[0]) / 2),
                        ceil((i[3] + i[1]) / 2)] for i in bbox]
            frame = np.array(cv2.imread(frame))
            for a, b in zip(image_ids, centers):
                image = cv2.putText(frame, a, (b[0] + 20, b[1] - 20), self.font, self.fontSize, 
                                    self.fontColor, self.thickness)
            cv2.imwrite(os.path.join(self.outputDir, imageName), image)
            self.image_counter += 1

class createExcel:
    def __init__(self):
        self.outputDir = os.path.join(os.getcwd(), "output")
        self.dateFormat = "%Y:%m:%d %H:%M:%S"
        self.seasons = {'spring': range(80, 172),
                        'summer': range(172, 264),
                        'fall': range(264, 355)}

    def output(self, exifData):
        # Create label of season sample time from image date
        imageDate = datetime.strptime(exifData['dateTime'], self.dateFormat)
        year = imageDate.year
        julianDay = imageDate.timetuple().tm_yday
        if julianDay in self.seasons['spring']:
            season = 'spring'
        elif julianDay in self.seasons['summer']:
            season = 'summer'
        elif julianDay in self.seasons['fall']:
            season = 'fall'
        else:
            season = 'winter'

        # Create Filename and blank dataframe for pandas excel formatting
        excelFilename = str(season) + '_' + str(year) + '_SWUAVSurvey' + '.xlsx'
        df = []

        # Convert all json files to pandas dataframe and export to temp excel file
        for subdir, _, files in os.walk(self.outputDir):
            for file in files:
                if file.endswith('.json'):
                    filename = os.path.join(subdir, file)
                    data = pd.read_json(filename, lines=True)
                    df.append(data)
        finalDF = pd.concat(df, ignore_index=True)
        finalDF.to_excel(os.path.join('output', excelFilename), index=False)

        # Reload excel file to add data validation (drop-down menu) to species column
        # based on list of likely waterfowl in Summit Lake and then save
        wb = load_workbook(os.path.join('output', excelFilename))
        wbNames = load_workbook('src/summitWaterflowList.xlsx').active
        wbNames._parent = wb
        wbNames.title = 'birdNames'
        wb._add_sheet(wbNames)
        wb.save(os.path.join('output', excelFilename))
        ws = wb.active
        dv = DataValidation(type = "list", formula1='birdNames!$A$1:$A$93',
                             allow_blank=True)
        ws.add_data_validation(dv)
        for row in range(2, ws.max_row + 1):
            dv.add(ws[f'E{row}'])
        wb.save(os.path.join('output', excelFilename))
                    
                    
